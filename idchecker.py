import pymongo, json, unicodedata, json
from pymongo import MongoClient
from openpyxl import load_workbook, Workbook

# connection = MongoClient('localhost', 27017)

# db = connection['wicket-wiz-db']
# document = db['test']

# post = {"_id":0, "name": "foo", "age":50}

# document.insert_one(post)

connection = MongoClient('localhost', 27017)

db = connection['cricmanagerrecent']
document = db['infoRowsCPL']

fetch = document.find()

def split(a, n):
    k, m = divmod(len(a), n)
    return (a[i*k+min(i, m):(i+1)*k+min(i+1, m)] for i in range(n))


mainarr = []
c = 0
for i in fetch:
	mainarr.append(i['_data'][0][11])

document2 = db['matchTrackerCPL']
for n in mainarr:
	post = {"_id": n}
	document2.insert_one(post)


# arrmainarr = split(mainarr, 7)

# for info in mainarr:
# 	print("added " + str(c))
# 	page.append(info)
# 	c += 1
# workbook = 'info.xlsx'
# wc_ = 0
# for m in arrmainarr:
# 	print('Chunk ' + str(wc_))
# 	wb = Workbook()

# 	ws = wb.create_sheet()
# 	r = 0
# 	for info in m:
# 		print(r)
# 		ws.append(info)
# 		r += 1
# 	wb.save(f"odimissing/odimissingpartone_{str(wc_)}.xlsx")
# 	wc_ += 1




