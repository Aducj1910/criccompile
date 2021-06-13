import pymongo, json, unicodedata, json, accessDB
from math import ceil
from pymongo import MongoClient
from openpyxl import load_workbook, Workbook

# connection = MongoClient('localhost', 27017)

# db = connection['wicket-wiz-db']
# document = db['test']

# post = {"_id":0, "name": "foo", "age":50}

# document.insert_one(post)

# connection = MongoClient('localhost', 27017)
connection = MongoClient('mongodb+srv://admin:21441@cluster0.7y5hx.mongodb.net/myFirstDatabase?retryWrites=true&w=majority')
db = connection['cricmanagerrecent']

inforows = 'infoRowsBBL'

document = db[inforows]

fetch = document.find()


def split(a, n):
    k, m = divmod(len(a), n)
    return (a[i*k+min(i, m):(i+1)*k+min(i+1, m)] for i in range(n))


mainarr = []
c = 0
for i in fetch:
	mainarr += i['_data']

# for i in fetch:
# 	accessDB.addNewMatch(i['_data'][0][11], matchtracker)
# 	print(i['_data'][0][11])


arrmainarr = split(mainarr, 1)

# for info in mainarr:
# 	print("added " + str(c))
# 	page.append(info)
# 	c += 1
workbook = 'info.xlsx'
wc_ = 0
for m in arrmainarr:
	print('Chunk ' + str(wc_))
	wb = Workbook()

	ws = wb.create_sheet()
	r = 0
	for info in m:
		print(r)
		ws.append(info)
		r += 1
	wb.save(f"odiupdate/{inforows}{str(wc_)}.xlsx")
	wc_ += 1




#RUN CRICSHEET FOR TEST, ODI, T20i, and others (folder is named)

#check player total after each match from scorecard to verify (using battracker, bowltracker)

#Bravo-Bravo repeat